"""xlsx → 每行一个 Chunk。"""
from __future__ import annotations
from app.core.types import Chunk, Modality, SourceType
from app.core.decorator import register
from app.core.exceptions import PreprocessError
from app.preprocess.base import BasePreprocessor
from app.utils.hashing import sha256_file, make_chunk_id
from app.utils.io import ensure_absolute


@register((".xlsx",))
class XlsxLoader(BasePreprocessor):
    """仅支持 xlsx（openpyxl 不支持 xls，xls 会抛 PreprocessError）。"""
    SUPPORTED_EXT = (".xlsx",)

    def load(self, file_path: str) -> list[Chunk]:
        path = ensure_absolute(file_path)
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise PreprocessError("缺少 openpyxl 依赖") from e

        try:
            wb = load_workbook(path, data_only=True, read_only=True)
        except Exception as e:
            raise PreprocessError(f"解析 xlsx 失败: {path}: {e}") from e

        file_hash = sha256_file(path)
        chunks: list[Chunk] = []
        idx = 0

        for sheet in wb.worksheets:
            sheet_name = sheet.title
            # 取第一行作为表头
            rows_iter = sheet.iter_rows(values_only=True)
            try:
                header = next(rows_iter)
            except StopIteration:
                continue
            header = [str(h).strip() if h is not None else "" for h in header]

            for r_idx, row in enumerate(rows_iter, start=2):
                parts: list[str] = []
                has_any = False
                for h, v in zip(header, row):
                    if v is None or v == "":
                        continue
                    has_any = True
                    if h:
                        parts.append(f"{h}: {v}")
                    else:
                        parts.append(str(v))
                if not has_any:
                    continue
                line = " | ".join(parts)
                chunks.append(Chunk(
                    chunk_id=make_chunk_id(path, idx),
                    source_path=path,
                    source_type=SourceType.XLSX,
                    modality=Modality.TEXT,
                    file_hash=file_hash,
                    sheet_name=sheet_name,
                    text=line,
                    preview=line[:200],
                    extra={"row": r_idx},
                ))
                idx += 1

        wb.close()
        return chunks
